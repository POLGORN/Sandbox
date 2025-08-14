import xml.etree.ElementTree as ET
import argparse
import logging
from openpyxl import load_workbook
from abc import ABC, abstractmethod
# ____________________________________________________________________
class XmlElement:
    """Свойства XML файла"""
    def __init__(self, tag, attributes, text):
        self.tag = tag
        self.attributes = attributes
        self.text = text


class XlsxRow:
    """Свойства XLSX файла"""
    def __init__(self, values):
        self.values = values


class DataReader(ABC):
    @abstractmethod
    def read(self, file_path):
        pass


class XmlReader(DataReader):
    """Метод для чтения XML"""
    def read(self, file_path):
        try:
            logging.info(f"Чтение XML файла: {file_path}...")
            tree = ET.parse(file_path)
            root = tree.getroot()
            elements = []

            for child in root:
                text_content = ''.join([elem.text.strip() for elem in child.iter() if elem.text is not None])
                element = XmlElement(child.tag, child.attrib, text_content)
                elements.append(element)


            logging.info(f"Успешно прочитано {len(elements)} элементов из XML")
            return elements
        except Exception as e:
            logging.error(f"Ошибка при чтении XML: {e}")
            return []


class XlsxReader(DataReader):
    """Метод для чтения XLSX"""
    def read(self, file_path):
        try:
            logging.info(f"Чтение XLSX файла: {file_path}...")
            workbook = load_workbook(file_path)
            sheet = workbook.active
            rows = []

            for row in sheet.iter_rows(values_only=True):
                rows.append(XlsxRow(row))

            logging.info(f"Успешно прочитано {len(rows)} элементов из XLSX")
            return rows
        except Exception as e:
            logging.error(f"Ошибка при чтении XLSX: {e}")
            return []
# ____________________________________________________________________
class DataView(ABC):
    @abstractmethod
    def display(self, data):
        pass


class XmlDataView(DataView):
    """Представление для XML"""
    def display(self, elements):
        print("\n--- XML Elements ---")
        for element in elements:
            print(f"Тег: {element.tag}, Атрибут: {element.attributes}, Значение: {element.text}")


class XlsxDataView(DataView):
    """Представление для XLSX"""
    def display(self, rows):
        print("\n--- XLSX Rows ---")
        for row in rows:
            print(f"Значение: {row.values}")
# ____________________________________________________________________
class DataProcessor:
    """Метод для инъекции зависимостей"""
    def __init__(self, reader: DataReader, view: DataView):
        self.reader = reader
        self.view = view

    """Метод контроля работы"""
    def process(self, file_path):
        data = self.reader.read(file_path)
        if data:
            self.view.display(data)
        else:
            logging.warning("Нет данных для отображения")
# ____________________________________________________________________
if __name__ == "__main__":
    """Настройка логирования"""
    logging.basicConfig(level=logging.INFO, format='>>> %(levelname)s - %(message)s')
    

    """Получение аргументов из строки терминала"""
    parser = argparse.ArgumentParser(description="Вывод данных из XML и XLSX файлов")
    parser.add_argument('--format', choices=['xml', 'xlsx', 'both'], default='both', 
                        help="Формат вывода: 'xml', 'xlsx' или 'both'")
    args = parser.parse_args()


    """Пути к данным"""
    xml_file_path = "input/data.xml"
    xlsx_file_path = "input/data.xlsx"


    """Создание экземпляров"""
    xml_reader = XmlReader()
    xlsx_reader = XlsxReader()


    xml_view = XmlDataView()
    xlsx_view = XlsxDataView()


    """Запуск программы"""
    try:
        if args.format in ['xml', 'both']:
            logging.info("Начало обработки XML файла...")
            processor = DataProcessor(xml_reader, xml_view)
            processor.process(xml_file_path)
            logging.info("Обработка XML файла завершена")
        if args.format in ['xlsx', 'both']:
            logging.info("Начало обработки XLSX файла...")
            processor = DataProcessor(xlsx_reader, xlsx_view)
            processor.process(xlsx_file_path)
            logging.info("Обработка XLSX файла завершена")
    except Exception as e:
        logging.error(f"Произошла ошибка в точке входа: {e}")