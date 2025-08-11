import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# Model
class XmlElement:
    def __init__(self, tag, attributes, text):
        self.tag = tag
        self.attributes = attributes
        self.text = text

class XlsxRow:
    def __init__(self, values):
        self.values = values

# Service
class DataReader:
    def __init__(self, file_path):
        self.file_path = file_path

    def read(self):
        if self.file_path.endswith('.xml'):
            return self.read_xml()
        elif self.file_path.endswith('.xlsx'):
            return self.read_xlsx()
        else:
            raise ValueError("Unsupported file format")

    def read_xml(self):
        tree = ET.parse(self.file_path)
        root = tree.getroot()
        elements = []

        for child in root:
            element = XmlElement(child.tag, child.attrib, child.text)
            elements.append(element)

        return elements

    def read_xlsx(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active
        rows = []

        for row in sheet.iter_rows(values_only=True):
            rows.append(XlsxRow(row))

        return rows

# View
class DataView:
    @staticmethod
    def display_xml_element(element):
        print(f"Тег: {element.tag}, Атрибуты: {element.attributes}, Значение: {element.text}")

    @staticmethod
    def display_xml_elements(elements):
        for element in elements:
            DataView.display_xml_element(element)

    @staticmethod
    def display_xlsx_row(row):
        print(f"Значения: {row.values}")

    @staticmethod
    def display_xlsx_rows(rows):
        for row in rows:
            DataView.display_xlsx_row(row)

# Main application
if __name__ == "__main__":
    file_path = "example.xlsx"  # Укажите путь к вашему XML или XLSX файлу
    reader = DataReader(file_path)
    view = DataView()

    # Чтение и отображение данных
    try:
        if file_path.endswith('.xml'):
            elements = reader.read()
            view.display_xml_elements(elements)
        elif file_path.endswith('.xlsx'):
            rows = reader.read()
            view.display_xlsx_rows(rows)
    except ValueError as e:
        print(e)
