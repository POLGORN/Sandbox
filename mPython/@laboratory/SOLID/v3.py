import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# Model
class XmlElement:
    def __init__(self, tag, attributes, text):
        self.tag = tag
        self.attributes = attributes
        self.text = text

class XlsxRow:
    def __init__(self, values):
        self.values = values

# Service
class DataReader:
    def read_xml(self, file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        elements = []

        for child in root:
            element = XmlElement(child.tag, child.attrib, child.text)
            elements.append(element)

        return elements

    def read_xlsx(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        rows = []

        for row in sheet.iter_rows(values_only=True):
            rows.append(XlsxRow(row))

        return rows

# View
class DataView:
    @staticmethod
    def display_xml_element(element):
        print(f"Тег: {element.tag}, Атрибуты: {element.attributes}, Значение: {element.text}")

    @staticmethod
    def display_xml_elements(elements):
        print("\n--- XML Elements ---")
        for element in elements:
            DataView.display_xml_element(element)

    @staticmethod
    def display_xlsx_row(row):
        print(f"Значения: {row.values}")

    @staticmethod
    def display_xlsx_rows(rows):
        print("\n--- XLSX Rows ---")
        for row in rows:
            DataView.display_xlsx_row(row)

# Main application
if __name__ == "__main__":
    xml_file_path = "example.xml"  # Укажите путь к вашему XML файлу
    xlsx_file_path = "example.xlsx"  # Укажите путь к вашему XLSX файлу

    reader = DataReader()
    view = DataView(
