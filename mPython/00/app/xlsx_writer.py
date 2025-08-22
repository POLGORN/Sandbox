from typing import Dict, List
from openpyxl import Workbook

class XlsxWriter:
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()

    def write_sheet(self, sheet_name: str, data: Dict[str, List[str]]):
        if not self.wb.sheetnames:
            ws = self.wb.active
            ws.title = sheet_name
        else:
            ws = self.wb.create_sheet(title=sheet_name)
        headers = list(data.keys())
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        max_rows = max((len(v) for v in data.values()), default=0)
        for r in range(max_rows):
            for c_idx, h in enumerate(headers, start=1):
                vals = data[h]
                val = vals[r] if r < len(vals) else ''
                ws.cell(row=r+2, column=c_idx, value=val)

    def save(self):
        self.wb.save(self.filename)