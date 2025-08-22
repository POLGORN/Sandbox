from abc import ABC, abstractmethod
from typing import Dict, List
import re
import sys
from openpyxl import Workbook

# ----------------------------
# Интерфейс парсера (IParser)
# ----------------------------
class IParser(ABC):
    @abstractmethod
    def parse(self, text: str) -> Dict[str, List[str]]:
        """Парсит текст и возвращает словарь column_name -> list_of_values"""
        ...

# ----------------------------
# Реализации парсеров
# ----------------------------
class Format1Parser(IParser):
    # Формат:
    # A: 1, 2, 3
    # B: 4, 5, 6
    # C: 7, 8, 9
    def parse(self, text: str) -> Dict[str, List[str]]:
        data: Dict[str, List[str]] = {}
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            m = re.match(r'^([^:]+)\s*:\s*(.+)$', line)
            if not m:
                continue
            key = m.group(1).strip()
            vals = [v.strip() for v in m.group(2).split(',') if v.strip() != '']
            data[key] = vals
        return data

class Format2Parser(IParser):
    # Формат CSV-подобный (первая строка — заголовки):
    # A, B, C
    # 1, 4, 7
    # 2, 5, 8
    # 3, 6, 9
    def parse(self, text: str) -> Dict[str, List[str]]:
        lines = [l.strip() for l in text.splitlines() if l.strip() != '']
        if not lines:
            return {}
        rows = [[c.strip() for c in re.split(r',\s*', line)] for line in lines]
        headers = rows[0]
        cols: Dict[str, List[str]] = {h: [] for h in headers}
        for row in rows[1:]:
            # расширяем до длины заголовков
            while len(row) < len(headers):
                row.append('')
            for i, h in enumerate(headers):
                cols[h].append(row[i] if i < len(row) else '')
        return cols

class Format3Parser(IParser):
    # Markdown таблица:
    # | A | B | C |
    # |---|---|---|
    # | 1 | 4 | 7 |
    # | 2 | 5 | 8 |
    # | 3 | 6 | 9 |
    def parse(self, text: str) -> Dict[str, List[str]]:
        lines = [l.strip() for l in text.splitlines() if l.strip() != '']
        if len(lines) < 2:
            return {}
        def split_row(line: str) -> List[str]:
            return [p.strip() for p in line.strip('|').split('|')]
        headers = split_row(lines[0])
        # пропускаем вторую строку (разделитель)
        data_rows = []
        for line in lines[2:]:
            data_rows.append(split_row(line))
        cols: Dict[str, List[str]] = {h: [] for h in headers}
        for row in data_rows:
            while len(row) < len(headers):
                row.append('')
            for i, h in enumerate(headers):
                cols[h].append(row[i])
        return cols

# ----------------------------
# Класс записи в XLSX
# ----------------------------
class XlsxWriter:
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()

    def write_sheet(self, sheet_name: str, data: Dict[str, List[str]]):
        # создаём лист; для первого листа используем активный
        if not self.wb.sheetnames:
            ws = self.wb.active
            ws.title = sheet_name
        else:
            ws = self.wb.create_sheet(title=sheet_name)
        headers = list(data.keys())
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        max_rows = max((len(v) for v in data.values()), default=0)
        for r in range(max_rows):
            for c_idx, h in enumerate(headers, start=1):
                vals = data[h]
                val = vals[r] if r < len(vals) else ''
                ws.cell(row=r+2, column=c_idx, value=val)

    def save(self):
        self.wb.save(self.filename)

# ----------------------------
# Приложение (Application Service)
# ----------------------------
class App:
    def __init__(self, parser1: IParser, parser2: IParser, parser3: IParser, writer: XlsxWriter):
        self.parser1 = parser1
        self.parser2 = parser2
        self.parser3 = parser3
        self.writer = writer

    def load_text(self, path: str) -> str:
        with open(path, 'r', encoding='utf-8') as f:
            return f.read()

    def run(self, in1_path: str, in2_path: str, in3_path: str, out_xlsx: str):
        t1 = self.load_text(in1_path)
        t2 = self.load_text(in2_path)
        t3 = self.load_text(in3_path)

        d1 = self.parser1.parse(t1)
        d2 = self.parser2.parse(t2)
        d3 = self.parser3.parse(t3)

        self.writer.filename = out_xlsx
        self.writer.write_sheet("Format1", d1)
        self.writer.write_sheet("Format2", d2)
        self.writer.write_sheet("Format3", d3)
        self.writer.save()

# ----------------------------
# CLI
# ----------------------------
def main(argv):
    if len(argv) != 5:
        print("Usage: python convert_to_xlsx.py input1.txt input2.txt input3.txt output.xlsx")
        return 1
    _, in1, in2, in3, out = argv
    app = App(Format1Parser(), Format2Parser(), Format3Parser(), XlsxWriter(out))
    app.run(in1, in2, in3, out)
    print("Saved:", out)
    return 0

if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
